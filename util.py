import json

import matplotlib.pyplot as plt
import numpy as np
import openpyxl as xl


# -----------------------------------------------------------------------------
# Linear Physical Programming
# -----------------------------------------------------------------------------


class LinearPhysicalProgramming:

    def __init__(self, problems):
        self.nps = len(problems)
        self.id = {'SIBS': list(), 'LIBS': list(), 'CIBS': list(), 'SIBH': list(), 'LIBH': list(), 'CIBH': list()}
        self.px = {'SIB': np.zeros((self.nps, 5)), 'LIB': np.zeros((self.nps, 5))}
        self.ds = {'SIB': np.zeros((self.nps, 5)), 'LIB': np.zeros((self.nps, 5))}
        for i, p in enumerate(problems):
            pclass = p['class'].upper()
            if pclass == 'SIBS':
                self.id['SIBS'].append(i)
                self.px['SIB'][i, :] = np.sort(np.array(p['metric']))
            elif pclass == 'LIBS':
                self.id['LIBS'].append(i)
                self.px['LIB'][i, :] = np.sort(np.array(p['metric']))[::-1]
            elif pclass == 'CIBS':
                self.id['CIBS'].append(i)
                ms = np.sort(np.array(p['metric']))
                self.px['LIB'][i, :] = ms[4::-1]
                self.px['SIB'][i, :] = ms[-5:]  # LIBS, SIBS
            elif pclass == 'SIBH':
                self.id['SIBH'].append(i)
                self.px['SIB'][i, 4] = p['metric']
            elif pclass == 'LIBH':
                self.id['LIBH'].append(i)
                self.px['LIB'][i, 4] = p['metric']
            elif pclass == 'CIBH':
                self.id['CIBH'].append(i)
                self.px['LIB'][i, 4] = np.min(np.array(p['metric']))
                self.px['SIB'][i, 4] = np.max(np.array(p['metric']))

        nsc = len(self.id['SIBS']) + len(self.id['LIBS']) + 2 * len(self.id['CIBS'])
        dy = np.zeros(5)
        dx = np.zeros(5)
        s = np.zeros(5)

        dy[1] = 0.1
        beta = 1.1
        dsmin = -np.inf
        while dsmin < 0.0:
            for k in range(2, 5):
                dy[k] = beta * nsc * dy[k - 1]
            for i in self.id['SIBS'] + self.id['CIBS']:
                dx[1:5] = self.px['SIB'][i, 1:5] - self.px['SIB'][i, 0:4]
                s[1:5] = dy[1:5] / dx[1:5]
                self.ds['SIB'][i, 1:5] = s[1:5] - s[0:4]
            for i in self.id['LIBS'] + self.id['CIBS']:
                dx[1:5] = -(self.px['LIB'][i, 1:5] - self.px['LIB'][i, 0:4])
                s[1:5] = dy[1:5] / dx[1:5]
                self.ds['LIB'][i, 1:5] = s[1:5] - s[0:4]
            dsmin = np.min((self.ds['SIB'], self.ds['LIB']))
            beta += 0.1
        self.py = np.cumsum(dy)

    @staticmethod
    def _calcSIBS(x, px, ds, ymax):
        p = np.searchsorted(px, x)
        if (p > 0) and (p < 5):
            d = x - px[0:p]
            y = np.sum(ds[1:(p + 1)] * d)
        elif p == 5:
            y = ymax
        else:
            y = 0.0
        return y, p

    def eval(self, x, ymax=np.inf):
        y = np.zeros(self.nps)
        p = np.zeros(self.nps)

        for i in self.id['SIBS']:
            y[i], p[i] = self._calcSIBS(x[i], self.px['SIB'][i, :], self.ds['SIB'][i, :], ymax)
        for i in self.id['LIBS']:
            y[i], p[i] = self._calcSIBS(-x[i], -self.px['LIB'][i, :], self.ds['LIB'][i, :], ymax)
        for i in self.id['CIBS']:
            if x[i] > self.px['SIB'][i, 0]:
                y[i], p[i] = self._calcSIBS(x[i], self.px['SIB'][i, :], self.ds['SIB'][i, :], ymax)
            else:
                y[i], p[i] = self._calcSIBS(-x[i], -self.px['LIB'][i, :], self.ds['LIB'][i, :], ymax)
        for i in self.id['SIBH']:
            if x[i] > self.px['SIB'][i, 4]:
                y[i], p[i] = np.inf, 5
        for i in self.id['LIBH']:
            if x[i] < self.px['LIB'][i, 4]:
                y[i], p[i] = np.inf, 5
        for i in self.id['CIBH']:
            if (x[i] < self.px['LIB'][i, 4]) or (x[i] > self.px['SIB'][i, 4]):
                y[i], p[i] = np.inf, 5
        self.x, self.y, self.p = x, y, p
        return np.mean(y)

    def plot(self):
        plt.figure()
        for i in self.id['SIBS']:
            plt.plot(self.px['SIB'][i, :], self.py, '.-')
            plt.plot(self.x[i], self.y[i], 'o')
        for i in self.id['LIBS']:
            plt.plot(self.px['LIB'][i, :], self.py, '.-')
            plt.plot(self.x[i], self.y[i], 'o')
        for i in self.id['CIBS']:
            plt.plot(np.hstack((self.px['LIB'][i, ::-1], self.px['SIB'][i, :])), np.hstack((self.py[::-1], self.py)), '.-')
            plt.plot(self.x[i], self.y[i], 'o')
        plt.grid()
        plt.show()


# -----------------------------------------------------------------------------
# File read utilities
# -----------------------------------------------------------------------------


def readxl(filename, sheet_name, cell_range):
    wb = xl.load_workbook(filename, data_only=True, read_only=True)
    ws = wb[sheet_name]

    min_col, min_row, max_col, max_row = xl.utils.range_boundaries(cell_range)
    if not (min_row or max_row):
        min_row = ws.min_row
        max_row = ws.max_row
    if not (min_col or max_col):
        min_col = ws.min_column
        max_col = ws.max_column

    v = np.zeros((max_row - min_row + 1, max_col - min_col + 1))
    for i, row in enumerate(ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col)):
        for j, cell in enumerate(row):
            v[i, j] = cell.value

    return v


# -----------------------------------------------------------------------------
# Etc
# -----------------------------------------------------------------------------


def cons_coefficient_matrix(x, degree=1):
    A = np.zeros((x.size, degree + 1))
    for j in range(degree + 1):
        A[:, degree - j] = x.reshape((x.size,))**j

    return A


def slope(x, y, x_lower=-np.inf, x_upper=np.inf, y_lower=-np.inf, y_upper=np.inf):
    ii = (x_lower < x) & (x < x_upper) & (y_lower < y) & (y < y_upper)
    A = cons_coefficient_matrix(x[ii], degree=1)
    s = np.linalg.lstsq(A, y[ii].reshape((y[ii].size,)))[0][0]

    return s


def intercept(x, y, x_lower=-np.inf, x_upper=np.inf, y_lower=-np.inf, y_upper=np.inf, axis='y'):
    ii = (x_lower < x) & (x < x_upper) & (y_lower < y) & (y < y_upper)
    if axis[0].upper() == 'Y':
        A = cons_coefficient_matrix(x[ii], degree=1)
        c = np.linalg.lstsq(A, y[ii].reshape((y[ii].size,)))[0][1]
    else:
        A = cons_coefficient_matrix(y[ii], degree=1)
        c = np.linalg.lstsq(A, x[ii].reshape((x[ii].size,)))[0][1]

    return c


# -----------------------------------------------------------------------------


def r2x(r, xb):
    return [(xb[i][1] - xb[i][0]) * r[i] + xb[i][0] for i in range(len(r))]


def x2r(x, xb):
    return [(x[i] - xb[i][0]) / (xb[i][1] - xb[i][0]) for i in range(len(x))]


def make_rb(xb):
    return [(0.0, 1.0) for i in range(len(xb))]


# -----------------------------------------------------------------------------


def update_mean(prev_mean, nth, nth_value):
    nth_mean = prev_mean * (nth - 1) / nth + nth_value / nth
    return nth_mean


def randn(ds, mu=0.0, sigma=1.0):
    return sigma * np.random.randn(ds) + mu


# -----------------------------------------------------------------------------


def get_value(dic, key):
    for k in dic.keys():
        if k[:len(key)].upper() == key.upper():
            return dic[k]
    return None


# -----------------------------------------------------------------------------


def input_dict_or_filename(inp):
    if isinstance(inp, dict):
        return inp
    with open(inp, 'r') as f:
        return json.load(f)


# -----------------------------------------------------------------------------
# matplotlib.pyplot utilities: adjust_axis()
# -----------------------------------------------------------------------------


def _nice_number(x, is_round):
    expv = np.floor(np.log10(x))
    f = x / pow(10.0, expv)  # between 1 and 10
    #
    if is_round:
        if f < 1.5:
            nf = 1
        elif f < 3.0:
            nf = 2
        elif f < 7.0:
            nf = 5
        else:
            nf = 10
    else:
        if f <= 1.0:
            nf = 1
        elif f <= 2.0:
            nf = 2
        elif f <= 5.0:
            nf = 5
        else:
            nf = 10
    #
    return nf * pow(10.0, expv)


def _adjust_axis_core(mini, maxi, nticks_desired=5):
    rng = _nice_number(maxi - mini, 0)
    spacing = _nice_number(rng / (nticks_desired - 1), 1)
    #
    mini = np.floor(mini / spacing) * spacing
    maxi = np.ceil(maxi / spacing) * spacing
    ntick = int((maxi - mini) / spacing + 0.5) + 1
    #
    return mini, maxi, spacing, ntick


def adjust_axis(xmin=None, xmax=None, ymin=None, ymax=None, nticks_desired=5):
    xminl, yminl, xmaxl, ymaxl = np.inf, np.inf, -np.inf, -np.inf
    ax = plt.gca()
    for line in ax.lines:
        xdata = line.get_xdata()
        ydata = line.get_ydata()

        xxdata = xdata[np.isfinite(ydata)]
        yydata = ydata[np.isfinite(xdata)]

        xminl = min([xminl, min(xxdata)])
        yminl = min([yminl, min(yydata)])
        xmaxl = max([xmaxl, max(xxdata)])
        ymaxl = max([ymaxl, max(yydata)])
    #
    if xmin is None:
        xmin = xminl
    if xmax is None:
        xmax = xmaxl
    if ymin is None:
        ymin = yminl
    if ymax is None:
        ymax = ymaxl
    xaxis = _adjust_axis_core(xmin, xmax, nticks_desired)
    yaxis = _adjust_axis_core(ymin, ymax, nticks_desired)
    #
    ax.axis([xaxis[0], xaxis[1], yaxis[0], yaxis[1]])
    ax.axes.set_xticks([xaxis[0] + d * xaxis[2] for d in range(xaxis[3])])
    ax.axes.set_yticks([yaxis[0] + d * yaxis[2] for d in range(yaxis[3])])
